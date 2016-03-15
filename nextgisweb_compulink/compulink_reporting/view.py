# coding=utf-8
import json
import tempfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Style
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from os import path
from pyramid.httpexceptions import HTTPForbidden
from pyramid.response import Response, FileResponse
from pyramid.view import view_config
from sqlalchemy import func
from nextgisweb import DBSession
from sqlalchemy.orm import joinedload_all
from .model import ConstructionStatusReport
from nextgisweb.pyramid import viewargs
from nextgisweb.resource import DataScope, ResourceGroup
from nextgisweb.resource.model import ResourceACLRule, Resource
from nextgisweb_compulink.compulink_admin import get_regions_from_resource, get_districts_from_resource, \
    get_project_statuses
from nextgisweb_compulink.compulink_admin.model import FoclProject, FoclStruct, PROJECT_STATUS_DELIVERED, \
    ConstructObject
from nextgisweb_compulink.compulink_admin.view import get_region_name, get_district_name
from nextgisweb_compulink.compulink_reporting.utils import DateTimeJSONEncoder
from view_ucn import add_routes
from nextgisweb_compulink.compulink_site import view as compulink_site_view

CURR_PATH = path.dirname(path.abspath(__file__))
TEMPLATES_PATH = path.join(CURR_PATH, 'templates/')


def setup_pyramid(comp, config):
    config.add_route(
        'compulink.reporting.status_grid',
        '/compulink/reporting/grid') \
        .add_view(status_grid)

    config.add_route(
        'compulink.reporting.get_status_report',
        '/compulink/reporting/get_status_report',
        client=()) \
        .add_view(get_status_report)

    config.add_route(
        'compulink.reporting.export_status_report',
        '/compulink/reporting/export_status_report',
        client=()) \
        .add_view(export_status_report)

    config.add_route(
        'compulink.reporting.building_objects',
        '/compulink/reporting/resources/child',
        client=()) \
        .add_view(get_child_resx_by_parent)

    add_routes(config)


@viewargs(renderer='nextgisweb_compulink:compulink_reporting/template/status_grid.mako')
def status_grid(request):
    if request.user.keyname == 'guest':
        raise HTTPForbidden()

    # prepare dicts
    all_regions = get_regions_from_resource()
    all_districts = get_districts_from_resource()

    acc_reg_id, acc_dist_id = get_user_accessible_structs(request.user)
    regions = filter(lambda x: x['id'] in acc_reg_id, all_regions)
    districts = filter(lambda x: x['id'] in acc_dist_id, all_districts)

    regions.sort(key=lambda x: x['name'].lower())
    districts.sort(key=lambda x: x['name'].lower())

    return dict(
        show_header=True,
        request=request,
        regions=regions,
        districts=districts
    )


def get_status_report(request):
    if request.user.keyname == 'guest':
        raise HTTPForbidden()

    report_query = construct_query(request)

    # dicts
    regions = get_regions_from_resource(as_dict=True)
    districts = get_districts_from_resource(as_dict=True)
    statuses = get_project_statuses(as_dict=True)

    json_report = []
    num_line = 1
    for row in report_query.all():
        json_row = {
            'num_line': num_line,
            'focl_name': row.focl_name,
            'region': regions.get(row.region, row.region),
            'district': districts.get(row.district, row.district),
            'status': statuses.get(row.status, row.status),
            'subcontr_name': row.subcontr_name,
            'start_build_time': row.start_build_time,
            'end_build_time': row.end_build_time,
            'cabling_plan': row.cabling_plan if row.cabling_plan else None,  # already in km
            'cabling_fact': row.cabling_fact if row.cabling_fact else None,  # already in km
            'cabling_percent': row.cabling_percent,
            'fosc_plan': row.fosc_plan,
            'fosc_fact': row.fosc_fact,
            'fosc_percent': row.fosc_percent,
            'cross_plan': row.cross_plan,
            'cross_fact': row.cross_fact,
            'cross_percent': row.cross_percent,
            'spec_trans_plan': row.spec_trans_plan,
            'spec_trans_fact': row.spec_trans_fact,
            'spec_trans_percent': row.spec_trans_percent,
            'ap_plan': row.ap_plan,
            'ap_fact': row.ap_fact,
            'ap_percent': row.ap_percent,
            'is_overdue': row.is_overdue,
            'is_month_overdue': row.is_month_overdue,
            'is_focl_delivered': row.status == PROJECT_STATUS_DELIVERED,
        }
        json_report.append(json_row)
        num_line += 1

    return Response(
        json.dumps(json_report, cls=DateTimeJSONEncoder),
        content_type=b'application/json')


def export_status_report(request):
    """
    ИНДЕКСЫ ЗАВЯЗАНЫ НА ШАБЛОН!
    """

    if request.user.keyname == 'guest':
        raise HTTPForbidden()

    # open template
    template_path = path.join(TEMPLATES_PATH, 'status_report.xlsx')
    wb = load_workbook(template_path)

    with tempfile.NamedTemporaryFile(delete=True) as temp_file:
        ws = wb.active

        # write headers
        region_filter = request.params.get('region', None)
        if region_filter:
            region = get_region_name(region_filter)
        else:
            region = 'Все'

        district_filter = request.params.get('district', None)
        if district_filter:
            district = get_district_name(district_filter)
        else:
            district = 'Все'

        status_filter = request.params.getall('status')
        if status_filter:
            st_names = get_project_statuses(as_dict=True)

            if len(set(st_names.keys()).difference(set(status_filter))) == 0:
                statuses = 'Все'
            else:
                statuses = set(st_names[st] for st in status_filter)
                statuses = ', '.join(statuses)
        else:
            statuses = 'Нет'

        only_overdue = 'Да' if request.params.get('only_overdue', False) else 'Нет'

        ws.cell(row=3, column=7).value = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        ws.cell(row=4, column=7).value = request.user.display_name
        ws.cell(row=3, column=3).value = region
        ws.cell(row=4, column=3).value = district
        ws.cell(row=5, column=3).value = statuses
        ws.cell(row=6, column=3).value = only_overdue

        # set borders (bug in lib!)
        border_style = ws.cell(row=9, column=7).style  # normal bordered cell :(
        for header_row in ws.get_squared_range(1, 8, 23, 9):
            for header_cell in header_row:
                header_cell.style = border_style

        # styles
        bold_style = ws.cell(row=3, column=2).style

        red_font = border_style.font.copy(color='F46A6A')
        red_bold_font = bold_style.font.copy(color='F46A6A')
        green_font = border_style.font.copy(color='008600')

        dt_style = Style(
            number_format='DD.MM.YYYY',
            font=border_style.font
        )

        overdue_style = Style(
            font=red_font
        )

        dt_overdue_style = Style(
            number_format='DD.MM.YYYY',
            font=red_font
        )

        month_overdue_style = Style(
            font=red_bold_font
        )

        dt_month_overdue_style = Style(
            number_format='DD.MM.YYYY',
            font=red_bold_font
        )

        delivered_style = Style(
            font=green_font
        )

        dt_delivered_style = Style(
            number_format='DD.MM.YYYY',
            font=green_font
        )

        footer_style = ws.cell(row=3, column=2).style
        footer_style_percent = Style(
            font=footer_style.font,
            number_format=FORMAT_PERCENTAGE
        )

        # get data
        report_query = construct_query(request)

        # dicts
        regions = get_regions_from_resource(as_dict=True)
        districts = get_districts_from_resource(as_dict=True)
        statuses = get_project_statuses(as_dict=True)

        # totals
        totals = {
            9: 0,
            10: 0,
            12: 0,
            13: 0,
            15: 0,
            16: 0,
            18: 0,
            19: 0,
            21: 0,
            22: 0,
        }

        num_line = 1
        for row in report_query.all():
            line_in_ws = 9 + num_line

            ws.cell(row=line_in_ws, column=1).value = num_line
            ws.cell(row=line_in_ws, column=2).value = row.focl_name
            ws.cell(row=line_in_ws, column=3).value = regions.get(row.region, row.region)
            ws.cell(row=line_in_ws, column=4).value = districts.get(row.district, row.district)
            ws.cell(row=line_in_ws, column=5).value = statuses.get(row.status, row.status)
            ws.cell(row=line_in_ws, column=6).value = row.subcontr_name
            ws.cell(row=line_in_ws, column=7).value = row.start_build_time
            ws.cell(row=line_in_ws, column=8).value = row.end_build_time
            ws.cell(row=line_in_ws, column=9).value = row.cabling_plan if row.cabling_plan else None
            ws.cell(row=line_in_ws, column=10).value = row.cabling_fact if row.cabling_fact else None
            ws.cell(row=line_in_ws, column=11).value = row.cabling_percent
            ws.cell(row=line_in_ws, column=12).value = row.fosc_plan
            ws.cell(row=line_in_ws, column=13).value = row.fosc_fact
            ws.cell(row=line_in_ws, column=14).value = row.fosc_percent
            ws.cell(row=line_in_ws, column=15).value = row.cross_plan
            ws.cell(row=line_in_ws, column=16).value = row.cross_fact
            ws.cell(row=line_in_ws, column=17).value = row.cross_percent
            ws.cell(row=line_in_ws, column=18).value = row.spec_trans_plan
            ws.cell(row=line_in_ws, column=19).value = row.spec_trans_fact
            ws.cell(row=line_in_ws, column=20).value = row.spec_trans_percent
            ws.cell(row=line_in_ws, column=21).value = row.ap_plan
            ws.cell(row=line_in_ws, column=22).value = row.ap_fact
            ws.cell(row=line_in_ws, column=23).value = row.ap_percent
            # special format fo DT
            ws.cell(row=line_in_ws, column=7).style = dt_style
            ws.cell(row=line_in_ws, column=8).style = dt_style

            if row.is_overdue:
                for col in xrange(1, 23 + 1):
                    ws.cell(row=line_in_ws, column=col).style = overdue_style
                ws.cell(row=line_in_ws, column=7).style = dt_overdue_style
                ws.cell(row=line_in_ws, column=8).style = dt_overdue_style

            if row.is_month_overdue:
                for col in xrange(1, 23 + 1):
                    ws.cell(row=line_in_ws, column=col).style = month_overdue_style
                ws.cell(row=line_in_ws, column=7).style = dt_month_overdue_style
                ws.cell(row=line_in_ws, column=8).style = dt_month_overdue_style

            if row.status == PROJECT_STATUS_DELIVERED:
                for col in xrange(1, 23 + 1):
                    ws.cell(row=line_in_ws, column=col).style = delivered_style
                ws.cell(row=line_in_ws, column=7).style = dt_delivered_style
                ws.cell(row=line_in_ws, column=8).style = dt_delivered_style

            # save totals
            for key in totals.keys():
                val = ws.cell(row=line_in_ws, column=key).value
                if val:
                    totals[key] += val

            num_line += 1

        # footer
        line_in_ws = 9 + num_line
        ws.merge_cells(start_row=line_in_ws, start_column=1, end_row=line_in_ws, end_column=6)
        ws.cell(row=line_in_ws, column=1).value = u'Итого'
        ws.cell(row=line_in_ws, column=1).style = footer_style

        for key in totals.keys():
            val = totals[key]
            if val:
                ws.cell(row=line_in_ws, column=key).value = val
                ws.cell(row=line_in_ws, column=key).style = footer_style

        for key in [11, 14, 17, 20, 23]:
            plan = ws.cell(row=line_in_ws, column=key - 2).value
            fact = ws.cell(row=line_in_ws, column=key - 1).value
            val = None
            if plan and fact:
                val = float(fact) / float(plan)
            if val:
                ws.cell(row=line_in_ws, column=key).value = val
                ws.cell(row=line_in_ws, column=key).style = footer_style_percent

        wb.save(path.abspath(temp_file.name))  # save to temp
        response = FileResponse(
            path.abspath(temp_file.name),
            content_type=bytes('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            request=request
        )
        response.content_disposition = u'attachment; filename="status report [%s].xlsx"' % \
                                       (datetime.now().strftime("%d_%m_%Y"))
        return response


def construct_query(request):
    # get params
    region_filter = request.params.get('region', None)
    district_filter = request.params.get('district', None)
    overdue_filter = request.params.get('only_overdue', False)
    status_filter = request.params.getall('status')
    project_filter = request.params.get('resource_id', None)

    report_query = DBSession.query(ConstructionStatusReport)
    if region_filter:
        report_query = report_query.filter(ConstructionStatusReport.region == region_filter)
    if district_filter:
        report_query = report_query.filter(ConstructionStatusReport.district == district_filter)
    if overdue_filter:
        report_query = report_query.filter(ConstructionStatusReport.is_overdue == True)

    report_query = report_query.filter(ConstructionStatusReport.status.in_(status_filter))

    if project_filter and project_filter != 'root':
        project_res_ids = get_project_focls(project_filter)
        report_query = report_query.filter(ConstructionStatusReport.focl_res_id.in_(project_res_ids))

    if not request.user.is_administrator:
        allowed_res_ids = get_user_writable_focls(request.user)
        report_query = report_query.filter(ConstructionStatusReport.focl_res_id.in_(allowed_res_ids))

    report_query = report_query.order_by(func.lower(ConstructionStatusReport.focl_name))

    return report_query

def get_project_focls(resource_id):
    try:
        root_res = DBSession.query(Resource).\
            filter(Resource.id == resource_id).\
            options(joinedload_all(Resource.children)).one()
    except:
        return []

    focl_ids = []
    def get_childs_recursive(resource):
        focl_ids.append(resource.id)
        if resource.identity in [ResourceGroup.identity, FoclProject.identity]:
            for child in resource.children:
                get_childs_recursive(child)
    get_childs_recursive(root_res)

    return focl_ids

def get_user_writable_focls(user):
    # get explicit rules
    rules_query = DBSession.query(ResourceACLRule) \
        .filter(ResourceACLRule.principal_id == user.principal_id) \
        .filter(ResourceACLRule.scope == DataScope.identity) \
        .options(joinedload_all(ResourceACLRule.resource))

    # todo: user groups explicit rules???
    # get permission for resource and children
    allowed_res_ids = []

    def get_perms_recursive(resource):
        # add self
        if resource.identity == FoclStruct.identity:
            if resource.has_permission(DataScope.write, user):
                allowed_res_ids.append(resource.id)
        elif resource.identity in [ResourceGroup.identity, FoclProject.identity]:
            allowed_res_ids.append(resource.id)
        # add childs
        if resource.identity in [ResourceGroup.identity, FoclProject.identity]:
            for child in resource.children:
                get_perms_recursive(child)

    for rule in rules_query.all():
        get_perms_recursive(rule.resource)

    return allowed_res_ids


def get_user_accessible_structs(user):
    query = DBSession.query(ConstructObject.region_id, ConstructObject.district_id)

    if not user.is_administrator:
        res_ids = get_user_writable_focls(user)
        query = query.filter(ConstructObject.resource_id.in_(res_ids))

    regions = []
    districts = []

    for region, district in query.all():
        if region not in regions:
            regions.append(region)
        if district not in districts:
            districts.append(district)

    return regions, districts


@view_config(renderer='json')
def get_child_resx_by_parent(request):
    response = compulink_site_view.get_child_resx_by_parent(request)

    if request.params.get('id', None) != '#':
        return response

    response_json = response.json

    for res_item in response_json:
        res_item['state'] = {'disabled': False}

    root_object = [{
        'res_type': 'focl_project',
        'text': u'Все',
        'has_children': True,
        'a_attr': {'chb': True},
        'children': response_json,
        'id': 'res_root',
        'icon': 'focl_project',
        'state': {
            'opened': True
        }
    }]

    return Response(json.dumps(root_object))
